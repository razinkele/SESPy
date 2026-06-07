import io

from openpyxl import load_workbook

from sespy.data_structure import Communication, Engagement, Stakeholder
from sespy.stakeholder_reports import (
    build_power_interest_png,
    build_stakeholder_workbook,
    build_summary_pdf,
)


def _id(k):  # fake translate for the PNG
    return k


def _fixture():
    sh = [
        Stakeholder(id="SH001", name="Port Authority", stakeholder_type="government",
                    power="HIGH", interest="HIGH"),
        Stakeholder(id="SH002", name="Coastal NGO", stakeholder_type="ngo",
                    power="LOW", interest="MEDIUM"),
    ]
    eng = [Engagement(id="ENG001", stakeholder_id="SH001", method="workshop")]
    comm = [Communication(id="COMM001", audience="key_players", comm_type="report")]
    return sh, eng, comm


def test_workbook_is_valid_xlsx_with_three_sheets():
    sh, eng, comm = _fixture()
    data = build_stakeholder_workbook(sh, eng, comm)
    assert data[:4] == b"PK\x03\x04"
    wb = load_workbook(io.BytesIO(data))
    assert wb.sheetnames == ["Stakeholders", "Engagements", "Communications"]
    ws = wb["Stakeholders"]
    assert ws.max_row == 3  # header + 2 data rows
    assert "Port Authority" in [c.value for c in ws[2]]


def test_workbook_empty_inputs_header_only():
    data = build_stakeholder_workbook([], [], [])
    wb = load_workbook(io.BytesIO(data))
    assert wb.sheetnames == ["Stakeholders", "Engagements", "Communications"]
    assert wb["Stakeholders"].max_row == 1  # header only


def test_png_is_valid_png():
    sh, _, _ = _fixture()
    data = build_power_interest_png(sh, translate=_id)
    assert data[:8] == b"\x89PNG\r\n\x1a\n"


def test_png_empty_inputs_still_valid():
    data = build_power_interest_png([], translate=_id)
    assert data[:8] == b"\x89PNG\r\n\x1a\n"


def test_pdf_is_valid_pdf():
    sh, eng, comm = _fixture()
    from sespy.stakeholders import stakeholder_stats
    stats = stakeholder_stats(sh, eng, comm)
    data = build_summary_pdf("My Project", stats, sh)
    assert data[:4] == b"%PDF"


def test_pdf_empty_inputs_still_valid():
    from sespy.stakeholders import stakeholder_stats
    data = build_summary_pdf("Empty", stakeholder_stats([], [], []), [])
    assert data[:4] == b"%PDF"


def test_pdf_escapes_markup_in_project_name():
    # reportlab Paragraph parses markup; an unescaped "<b>" would raise.
    from sespy.stakeholders import stakeholder_stats
    data = build_summary_pdf("<b>Bad & Co", stakeholder_stats([], [], []), [])
    assert data[:4] == b"%PDF"
